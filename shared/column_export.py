"""
Utilidades reutilizables para exportar listados a Excel/PDF respetando
las columnas visibles que el usuario eligió en el selector de columnas.

Usadas por las vistas que implementan su propio selector de columnas
(Marcas, Categorías, Proveedores, además del patrón ya existente en
Productos/Clientes). Centraliza el armado de Workbook/PDF para no repetir
~80 líneas de estilos por cada módulo nuevo.
"""
from django.http import HttpResponse
from django.utils import timezone


def export_visible_columns_excel(queryset, visible_columns, all_columns, get_value, title, filename_base):
    """Genera un .xlsx con solo las columnas visibles seleccionadas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    headers = [all_columns[col]['label'] for col in visible_columns]
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='343A40')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for obj in queryset:
        row = [str(get_value(obj, col)) for col in visible_columns]
        ws.append(row)

    widths = [len(h) for h in headers]
    for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for c, cell in enumerate(row):
            cell.border = border
            widths[c] = max(widths[c], len(str(cell.value or '')))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w + 4, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'{filename_base}_{timezone.localtime().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_visible_columns_pdf(queryset, visible_columns, all_columns, get_value, title, filename_base):
    """Genera un .pdf con solo las columnas visibles seleccionadas."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    response = HttpResponse(content_type='application/pdf')
    filename = f'{filename_base}_{timezone.localtime().strftime("%Y%m%d_%H%M")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pagesize = landscape(A4) if len(visible_columns) > 5 else A4
    doc = SimpleDocTemplate(
        response, pagesize=pagesize,
        leftMargin=0.8 * cm, rightMargin=0.8 * cm,
        topMargin=0.8 * cm, bottomMargin=0.8 * cm,
    )
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(title, styles['Title']),
        Paragraph(f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
        Spacer(1, 0.3 * cm),
    ]

    headers = [all_columns[col]['label'] for col in visible_columns]
    data = [headers]
    for obj in queryset:
        data.append([str(get_value(obj, col)) for col in visible_columns])

    font_size = 7 if len(visible_columns) > 8 else (8 if len(visible_columns) > 5 else 9)

    table = Table(data, colWidths=[(pagesize[0] - 1.6 * cm) / len(visible_columns)] * len(visible_columns))
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), font_size),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))

    elements.append(table)
    doc.build(elements)
    return response
