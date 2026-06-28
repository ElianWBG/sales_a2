from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum, Avg, Max, Min, Count
from django.db.models import F
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib.auth import login, get_user_model
from django.utils import timezone
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from datetime import timedelta, date
from .models import *
from .forms import (
    SignUpForm, BrandForm, ProductGroupForm, SupplierForm,
    CustomerForm, InvoiceForm, InvoiceDetailFormSet
)
from .ProductForm import ProductForm
from shared.export_mixins import ExportListMixin
from shared.mixins import StaffRequiredMixin
from shared.decorators import audit_action
from .column_config import get_visible_columns, get_all_columns, validate_visible_columns, DEFAULT_VISIBLE_COLUMNS
from .brand_column_config import (
    get_brand_visible_columns, get_all_brand_columns,
    validate_brand_visible_columns, BRAND_DEFAULT_VISIBLE_COLUMNS
)
from .productgroup_column_config import (
    get_productgroup_visible_columns, get_all_productgroup_columns,
    validate_productgroup_visible_columns, PRODUCTGROUP_DEFAULT_VISIBLE_COLUMNS
)
from .supplier_column_config import (
    get_supplier_visible_columns, get_all_supplier_columns,
    validate_supplier_visible_columns, SUPPLIER_DEFAULT_VISIBLE_COLUMNS
)
from shared.column_export import export_visible_columns_excel, export_visible_columns_pdf
from .customer_column_config import (
    get_customer_visible_columns, get_all_customer_columns,
    validate_customer_visible_columns, CUSTOMER_DEFAULT_VISIBLE_COLUMNS
)
from .invoice_column_config import (
    get_invoice_visible_columns, get_all_invoice_columns,
    validate_invoice_visible_columns, INVOICE_DEFAULT_VISIBLE_COLUMNS
)
from decimal import Decimal


# === HOME / DASHBOARD ===
@login_required
@audit_action('VIEW_HOME')
def home(request):
    """Panel principal: tarjetas resumen, gráficos, accesos rápidos,
    actividad reciente y alertas del sistema."""
    User = get_user_model()
    today = timezone.localdate()
    week_ago = today - timedelta(days=7)

    # --- Tarjetas resumen ---
    total_products = Product.objects.count()
    total_categories = ProductGroup.objects.count()
    total_brands = Brand.objects.count()
    total_suppliers = Supplier.objects.count()
    total_customers = Customer.objects.count()
    total_users = User.objects.count()
    total_sales = Invoice.objects.count()
    total_income = Invoice.objects.filter(is_active=True).aggregate(s=Sum('total'))['s'] or 0

    from purchasing.models import Purchase
    total_purchases = Purchase.objects.count()

    low_stock_qs = Product.objects.filter(is_active=True, stock__gt=0, stock__lte=10)
    out_of_stock_qs = Product.objects.filter(is_active=True, stock=0)
    low_stock_count = low_stock_qs.count()
    out_of_stock_count = out_of_stock_qs.count()

    # --- Gráfico: ventas por mes (últimos 6 meses) ---
    months = []
    cursor_year, cursor_month = today.year, today.month
    for i in range(5, -1, -1):
        y, m = cursor_year, cursor_month - i
        while m <= 0:
            m += 12
            y -= 1
        months.append(date(y, m, 1))

    sales_by_month = (
        Invoice.objects
        .filter(invoice_date__date__gte=months[0])
        .annotate(month=TruncMonth('invoice_date'))
        .values('month')
        .annotate(total=Sum('total'), count=Count('id'))
    )
    sales_map = {row['month'].strftime('%Y-%m'): row for row in sales_by_month if row['month']}
    month_labels = [m.strftime('%b %Y') for m in months]
    month_totals = [float(sales_map.get(m.strftime('%Y-%m'), {}).get('total') or 0) for m in months]
    month_counts = [sales_map.get(m.strftime('%Y-%m'), {}).get('count') or 0 for m in months]

    # --- Gráfico: productos más vendidos ---
    top_products_qs = (
        InvoiceDetail.objects
        .values('product__name')
        .annotate(qty=Sum('quantity'))
        .order_by('-qty')[:5]
    )
    top_products_labels = [row['product__name'] for row in top_products_qs]
    top_products_data = [row['qty'] for row in top_products_qs]

    # --- Gráfico: distribución de categorías ---
    category_dist_qs = (
        ProductGroup.objects
        .annotate(num_products=Count('products'))
        .filter(num_products__gt=0)
        .order_by('-num_products')
    )
    category_labels = [g.name for g in category_dist_qs]
    category_data = [g.num_products for g in category_dist_qs]

    # --- Gráfico: estado del stock (sustituye a "Compras vs Ventas", aún no implementado) ---
    healthy_stock_count = Product.objects.filter(is_active=True, stock__gt=10).count()
    stock_status_labels = ['Stock saludable', 'Stock bajo', 'Agotado']
    stock_status_data = [healthy_stock_count, low_stock_count, out_of_stock_count]

    # --- Actividad reciente ---
    recent_sales = Invoice.objects.select_related('customer').order_by('-invoice_date')[:5]
    recent_products = Product.objects.select_related('brand', 'group').order_by('-created_at')[:5]
    recent_users = User.objects.order_by('-date_joined')[:5]

    # --- Alertas del sistema ---
    sales_today = Invoice.objects.filter(invoice_date__date=today)
    sales_today_count = sales_today.count()
    sales_today_total = sales_today.aggregate(s=Sum('total'))['s'] or 0
    new_customers_week = Customer.objects.filter(created_at__date__gte=week_ago).count()

    context = {
        'total_products': total_products,
        'total_categories': total_categories,
        'total_brands': total_brands,
        'total_suppliers': total_suppliers,
        'total_customers': total_customers,
        'total_users': total_users,
        'total_sales': total_sales,
        'total_purchases': total_purchases,
        'total_income': total_income,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,

        'month_labels': month_labels,
        'month_totals': month_totals,
        'month_counts': month_counts,
        'top_products_labels': top_products_labels,
        'top_products_data': top_products_data,
        'category_labels': category_labels,
        'category_data': category_data,
        'stock_status_labels': stock_status_labels,
        'stock_status_data': stock_status_data,

        'recent_sales': recent_sales,
        'recent_products': recent_products,
        'recent_users': recent_users,

        'low_stock_products': low_stock_qs.order_by('stock')[:5],
        'out_of_stock_products': out_of_stock_qs[:5],
        'sales_today_count': sales_today_count,
        'sales_today_total': sales_today_total,
        'new_customers_week': new_customers_week,
    }
    return render(request, 'billing/home.html', context)


# === REGISTRO ===
class SignUpView(CreateView):
    form_class = SignUpForm
    template_name = 'registration/signup.html'
    success_url = reverse_lazy('billing:home')
    def form_valid(self, form):
        response = super().form_valid(form)
        login(self.request, self.object)
        return response

# === BRAND (FBV) ===
@login_required
@audit_action('LIST_BRANDS')
def brand_list(request):
    qs = Brand.objects.all()
    name = request.GET.get('name', '').strip()
    is_active = request.GET.get('is_active', '')
    if name:
        qs = qs.filter(name__icontains=name)
    if is_active in ('0', '1'):
        qs = qs.filter(is_active=is_active == '1')

    # Columnas visibles (selector de columnas)
    visible_columns_list = request.session.get('brand_visible_columns', BRAND_DEFAULT_VISIBLE_COLUMNS)
    is_valid, visible_columns_list = validate_brand_visible_columns(visible_columns_list)

    def get_export_value(obj, col_key):
        if col_key == 'name':
            return obj.name
        elif col_key == 'description':
            return obj.description or '-'
        elif col_key == 'product_count':
            return obj.products.count()
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        return getattr(obj, col_key, '-')

    # Export (respeta las columnas visibles seleccionadas)
    export = request.GET.get('export')
    if export in ('excel', 'pdf'):
        all_columns = get_all_brand_columns()
        if export == 'excel':
            return export_visible_columns_excel(qs, visible_columns_list, all_columns, get_export_value, 'Marcas', 'Marcas')
        return export_visible_columns_pdf(qs, visible_columns_list, all_columns, get_export_value, 'Listado de Marcas', 'Marcas')

    # Paginación manual
    from django.core.paginator import Paginator
    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    params = request.GET.copy()
    params.pop('page', None)

    return render(request, 'billing/brand_list.html', {
        'brands': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'paginator': paginator,
        'search_params': params.urlencode(),
        'visible_columns': get_brand_visible_columns(visible_columns_list),
        'all_columns': get_all_brand_columns(),
        'visible_columns_list': visible_columns_list,
    })

@login_required
def brand_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de marcas"""
    from django.http import JsonResponse
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])
        is_valid, validated_columns = validate_brand_visible_columns(visible_columns)
        request.session['brand_visible_columns'] = validated_columns
        request.session.modified = True
        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_brand_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@audit_action('CREATE_BRAND')
def brand_create(request):
    if request.method == 'POST':
        form = BrandForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marca creada exitosamente!')
            return redirect('billing:brand_list')
    else:
        form = BrandForm()
    return render(request, 'billing/brand_form.html', {'form': form, 'title': 'Crear Marca'})

@login_required
@audit_action('VIEW_BRAND')
def brand_detail(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    products = brand.products.all()[:10]
    return render(request, 'billing/brand_detail.html', {
        'brand': brand,
        'products': products,
        'product_count': brand.products.count(),
    })

@login_required
@audit_action('UPDATE_BRAND')
def brand_update(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        form = BrandForm(request.POST, instance=brand)
        if form.is_valid():
            form.save()
            messages.success(request, 'Marca actualizada exitosamente!')
            return redirect('billing:brand_list')
    else:
        form = BrandForm(instance=brand)
    return render(request, 'billing/brand_form.html', {'form': form, 'title': 'Editar Marca'})

@login_required
@audit_action('DELETE_BRAND')
def brand_delete(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        brand.delete()
        messages.success(request, 'Marca eliminada exitosamente!')
        return redirect('billing:brand_list')
    return render(request, 'billing/brand_confirm_delete.html', {'object': brand})


# === PRODUCTGROUP (CBV) ===
class ProductGroupListView(ExportListMixin, LoginRequiredMixin, ListView):
    model = ProductGroup
    template_name = 'billing/product_group_list.html'
    context_object_name = 'items'
    paginate_by = 10

    export_title = 'Grupos de Productos'
    export_fields = [
        ('Nombre', 'name'),
        ('Estado', lambda o: 'Activo' if o.is_active else 'Inactivo'),
        ('Creación', lambda o: o.created_at.strftime('%d/%m/%Y')),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt in ('excel', 'pdf'):
            visible_columns = self.get_visible_cols()
            all_columns = get_all_productgroup_columns()
            qs = self.get_queryset()
            if fmt == 'excel':
                return export_visible_columns_excel(qs, visible_columns, all_columns, self.get_export_value, 'Categorías', 'Categorias')
            return export_visible_columns_pdf(qs, visible_columns, all_columns, self.get_export_value, 'Listado de Categorías', 'Categorias')
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = ProductGroup.objects.all()
        g = self.request.GET
        if name := g.get('name', '').strip():
            qs = qs.filter(name__icontains=name)
        if (is_active := g.get('is_active', '')) in ('0', '1'):
            qs = qs.filter(is_active=is_active == '1')
        return qs

    def get_visible_cols(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('productgroup_visible_columns', PRODUCTGROUP_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_productgroup_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        if col_key == 'name':
            return obj.name
        elif col_key == 'product_count':
            return obj.products.count()
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        return getattr(obj, col_key, '-')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()

        visible_columns = self.get_visible_cols()
        ctx['visible_columns'] = get_productgroup_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_productgroup_columns()
        ctx['visible_columns_list'] = visible_columns
        return ctx

class ProductGroupCreateView(LoginRequiredMixin, CreateView):
    model = ProductGroup; form_class = ProductGroupForm
    template_name = 'billing/product_group_form.html'
    success_url = reverse_lazy('billing:productgroup_list')

class ProductGroupUpdateView(LoginRequiredMixin, UpdateView):
    model = ProductGroup; form_class = ProductGroupForm
    template_name = 'billing/product_group_form.html'
    success_url = reverse_lazy('billing:productgroup_list')

class ProductGroupDetailView(LoginRequiredMixin, DetailView):
    model = ProductGroup
    template_name = 'billing/product_group_detail.html'
    context_object_name = 'group'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['products'] = self.object.products.all()[:10]
        ctx['product_count'] = self.object.products.count()
        return ctx

class ProductGroupDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    model = ProductGroup
    template_name = 'billing/product_group_confirm_delete.html'
    success_url = reverse_lazy('billing:productgroup_list')
    staff_redirect_url = '/groups/'

@login_required
def productgroup_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de categorías"""
    from django.http import JsonResponse
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])
        is_valid, validated_columns = validate_productgroup_visible_columns(visible_columns)
        request.session['productgroup_visible_columns'] = validated_columns
        request.session.modified = True
        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_productgroup_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# === SUPPLIER (CBV) ===
class SupplierListView(ExportListMixin, LoginRequiredMixin, ListView):
    model = Supplier
    template_name = 'billing/supplier_list.html'
    context_object_name = 'items'
    paginate_by = 10

    export_title = 'Proveedores'
    export_fields = [
        ('Nombre', 'name'),
        ('Contacto', lambda o: o.contact_name or '-'),
        ('Email', lambda o: o.email or '-'),
        ('Teléfono', lambda o: o.phone or '-'),
        ('Estado', lambda o: 'Activo' if o.is_active else 'Inactivo'),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt in ('excel', 'pdf'):
            visible_columns = self.get_visible_cols()
            all_columns = get_all_supplier_columns()
            qs = self.get_queryset()
            if fmt == 'excel':
                return export_visible_columns_excel(qs, visible_columns, all_columns, self.get_export_value, 'Proveedores', 'Proveedores')
            return export_visible_columns_pdf(qs, visible_columns, all_columns, self.get_export_value, 'Listado de Proveedores', 'Proveedores')
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Supplier.objects.all()
        g = self.request.GET
        if name := g.get('name', '').strip():
            qs = qs.filter(name__icontains=name)
        if email := g.get('email', '').strip():
            qs = qs.filter(email__icontains=email)
        if (is_active := g.get('is_active', '')) in ('0', '1'):
            qs = qs.filter(is_active=is_active == '1')
        return qs

    def get_visible_cols(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('supplier_visible_columns', SUPPLIER_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_supplier_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        if col_key == 'name':
            return obj.name
        elif col_key == 'contact_name':
            return obj.contact_name or '-'
        elif col_key == 'email':
            return obj.email or '-'
        elif col_key == 'phone':
            return obj.phone or '-'
        elif col_key == 'address':
            return obj.address or '-'
        elif col_key == 'product_count':
            return obj.products.count()
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        return getattr(obj, col_key, '-')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()

        visible_columns = self.get_visible_cols()
        ctx['visible_columns'] = get_supplier_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_supplier_columns()
        ctx['visible_columns_list'] = visible_columns
        return ctx

class SupplierCreateView(LoginRequiredMixin, CreateView):
    model = Supplier; form_class = SupplierForm
    template_name = 'billing/supplier_form.html'
    success_url = reverse_lazy('billing:supplier_list')

class SupplierUpdateView(LoginRequiredMixin, UpdateView):
    model = Supplier; form_class = SupplierForm
    template_name = 'billing/supplier_form.html'
    success_url = reverse_lazy('billing:supplier_list')

class SupplierDetailView(LoginRequiredMixin, DetailView):
    model = Supplier
    template_name = 'billing/supplier_detail.html'
    context_object_name = 'supplier'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['products'] = self.object.products.all()[:10]
        ctx['product_count'] = self.object.products.count()
        return ctx

class SupplierDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    model = Supplier
    template_name = 'billing/supplier_confirm_delete.html'
    success_url = reverse_lazy('billing:supplier_list')
    staff_redirect_url = '/suppliers/'

@login_required
def supplier_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de proveedores"""
    from django.http import JsonResponse
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])
        is_valid, validated_columns = validate_supplier_visible_columns(visible_columns)
        request.session['supplier_visible_columns'] = validated_columns
        request.session.modified = True
        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_supplier_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# === PRODUCT (CBV) ===
class ProductListView(ExportListMixin, LoginRequiredMixin, ListView):
    model = Product
    template_name = 'billing/product_list.html'
    context_object_name = 'items'
    paginate_by = 3

    export_title = 'Productos'
    export_fields = [
        ('Nombre', 'name'),
        ('Marca', 'brand.name'),
        ('Grupo', 'group.name'),
        ('Precio', lambda o: f'{o.unit_price:.2f}'),
        ('Stock', 'stock'),
        ('Proveedores', lambda o: ', '.join(s.name for s in o.suppliers.all()) or '-'),
        ('Estado', lambda o: 'Activo' if o.is_active else 'Inactivo'),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt == 'excel':
            return self.export_excel_with_visible_columns()
        if fmt == 'pdf':
            return self.export_pdf_with_visible_columns()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Product.objects.select_related('brand', 'group').prefetch_related('suppliers')
        g = self.request.GET
        if name := g.get('name', '').strip():
            qs = qs.filter(name__icontains=name)
        if brand := g.get('brand', ''):
            qs = qs.filter(brand_id=brand)
        if group := g.get('group', ''):
            qs = qs.filter(group_id=group)
        if price_min := g.get('price_min', '').strip():
            qs = qs.filter(unit_price__gte=price_min)
        if price_max := g.get('price_max', '').strip():
            qs = qs.filter(unit_price__lte=price_max)
        if stock_min := g.get('stock_min', '').strip():
            qs = qs.filter(stock__gte=stock_min)
        if stock_max := g.get('stock_max', '').strip():
            qs = qs.filter(stock__lte=stock_max)
        if (is_active := g.get('is_active', '')) in ('0', '1'):
            qs = qs.filter(is_active=is_active == '1')
        if supplier := g.get('supplier', ''):
            qs = qs.filter(suppliers__id=supplier).distinct()
        return qs

    def get_visible_columns(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('product_visible_columns', DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        """Obtener el valor de una columna específica"""
        col_config = get_all_columns()[col_key]
        field = col_config['field']
        
        if col_key == 'image':
            return 'Sí' if obj.image else 'No'
        elif col_key == 'brand':
            return obj.brand.name if obj.brand else '-'
        elif col_key == 'group':
            return obj.group.name if obj.group else '-'
        elif col_key == 'suppliers':
            return ', '.join(s.name for s in obj.suppliers.all()) or '-'
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'unit_price':
            return f'{obj.unit_price:.2f}'
        elif col_key == 'balance':
            return f'{obj.balance:.2f}'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        else:
            return getattr(obj, field, '-')

    def export_excel_with_visible_columns(self):
        """Exportar a Excel con columnas visibles"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        visible_columns = self.get_visible_columns()
        all_columns = get_all_columns()
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Productos'[:31]
        
        # Encabezados
        headers = [all_columns[col]['label'] for col in visible_columns]
        ws.append(headers)
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='343A40')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        
        # Datos
        rows = self.get_queryset()
        for row_obj in rows:
            row = [str(self.get_export_value(row_obj, col)) for col in visible_columns]
            ws.append(row)
        
        # Ajustar ancho de columnas
        widths = [len(h) for h in headers]
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for c, cell in enumerate(row):
                cell.border = border
                widths[c] = max(widths[c], len(str(cell.value or '')))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w + 4, 60)
        
        # Respuesta
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Productos_{timezone.localtime().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def export_pdf_with_visible_columns(self):
        """Exportar a PDF con columnas visibles"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape, portrait
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from django.http import HttpResponse
        
        visible_columns = self.get_visible_columns()
        all_columns = get_all_columns()
        
        response = HttpResponse(content_type='application/pdf')
        filename = f'Productos_{timezone.localtime().strftime("%Y%m%d_%H%M")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Determinar orientación basada en número de columnas
        pagesize = landscape(A4) if len(visible_columns) > 5 else A4
        
        doc = SimpleDocTemplate(
            response, pagesize=pagesize,
            leftMargin=0.8 * cm, rightMargin=0.8 * cm,
            topMargin=0.8 * cm, bottomMargin=0.8 * cm,
        )
        
        styles = getSampleStyleSheet()
        
        # Ajustar tamaño de fuente según número de columnas
        if len(visible_columns) > 8:
            font_size = 7
        elif len(visible_columns) > 5:
            font_size = 8
        else:
            font_size = 9
        
        cell_style = ParagraphStyle('cell', parent=styles['Normal'], 
                                   fontSize=font_size, leading=font_size + 2)
        head_style = ParagraphStyle('cellHead', parent=styles['Normal'],
                                   fontSize=font_size, leading=font_size + 2,
                                   textColor=colors.white, fontName='Helvetica-Bold')
        
        elements = [
            Paragraph('Listado de Productos', styles['Title']),
            Paragraph(
                f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
                styles['Normal'],
            ),
            Spacer(1, 0.3 * cm),
        ]
        
        # Encabezados
        headers = [all_columns[col]['label'] for col in visible_columns]
        data = [headers]
        
        # Datos
        for obj in self.get_queryset():
            row = [str(self.get_export_value(obj, col)) for col in visible_columns]
            data.append(row)
        
        # Crear tabla
        table = Table(data, colWidths=[(pagesize[0] - 1.6 * cm) / len(visible_columns)] * len(visible_columns))
        
        # Estilos de tabla
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['brands'] = Brand.objects.order_by('name')
        ctx['groups'] = ProductGroup.objects.order_by('name')
        ctx['suppliers'] = Supplier.objects.order_by('name')
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()
        
        # Obtener columnas visibles de la sesión o usar default
        visible_columns = self.request.session.get('product_visible_columns', DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns = validate_visible_columns(visible_columns)
        
        ctx['visible_columns'] = get_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_columns()
        ctx['visible_columns_list'] = visible_columns
        
        return ctx

class ProductCreateView(LoginRequiredMixin, CreateView):
    model = Product; form_class = ProductForm
    template_name = 'billing/product_form.html'
    success_url = reverse_lazy('billing:product_list')

class ProductUpdateView(LoginRequiredMixin, UpdateView):
    model = Product; form_class = ProductForm
    template_name = 'billing/product_form.html'
    success_url = reverse_lazy('billing:product_list')

class ProductDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    model = Product
    template_name = 'billing/product_confirm_delete.html'
    success_url = reverse_lazy('billing:product_list')
    staff_redirect_url = '/products/'


class ProductDetailView(LoginRequiredMixin, DetailView):
    model = Product
    template_name = 'billing/product_detail.html'
    context_object_name = 'product'


@login_required
def product_update_image(request, pk):
    """Actualizar imagen del producto via AJAX"""
    from django.http import JsonResponse
    from django.views.decorators.http import require_http_methods
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    product = get_object_or_404(Product, pk=pk)
    
    if 'image' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No se envió imagen'}, status=400)
    
    image_file = request.FILES['image']
    
    # Validar tipo de archivo
    valid_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
    if image_file.content_type not in valid_types:
        return JsonResponse({'success': False, 'error': 'Tipo de archivo no válido'}, status=400)
    
    # Validar tamaño (5MB)
    if image_file.size > 5 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'Archivo muy grande (máx: 5MB)'}, status=400)
    
    try:
        product.image = image_file
        product.save()
        return JsonResponse({
            'success': True,
            'image_url': product.image.url,
            'message': 'Imagen actualizada correctamente'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def product_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de productos"""
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])
        
        # Validar columnas
        is_valid, validated_columns = validate_visible_columns(visible_columns)
        
        # Guardar en sesión
        request.session['product_visible_columns'] = validated_columns
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# === CUSTOMER (CBV) ===
class CustomerListView(ExportListMixin, LoginRequiredMixin, ListView):
    model = Customer
    template_name = 'billing/customer_list.html'
    context_object_name = 'items'
    paginate_by = 10

    export_title = 'Clientes'
    export_fields = [
        ('DNI/RUC', 'dni'),
        ('Apellido', 'last_name'),
        ('Nombre', 'first_name'),
        ('Email', lambda o: o.email or '-'),
        ('Teléfono', lambda o: o.phone or '-'),
        ('Estado', lambda o: 'Activo' if o.is_active else 'Inactivo'),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt == 'excel':
            return self.export_excel_with_visible_columns()
        if fmt == 'pdf':
            return self.export_pdf_with_visible_columns()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Customer.objects.all()
        g = self.request.GET
        if name := g.get('name', '').strip():
            qs = qs.filter(first_name__icontains=name) | qs.filter(last_name__icontains=name)
        if dni := g.get('dni', '').strip():
            qs = qs.filter(dni__icontains=dni)
        if email := g.get('email', '').strip():
            qs = qs.filter(email__icontains=email)
        if (is_active := g.get('is_active', '')) in ('0', '1'):
            qs = qs.filter(is_active=is_active == '1')
        return qs

    def get_customer_visible_cols(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('customer_visible_columns', CUSTOMER_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_customer_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        """Obtener el valor de una columna específica"""
        if col_key == 'dni':
            return obj.dni
        elif col_key == 'last_name':
            return obj.last_name
        elif col_key == 'first_name':
            return obj.first_name
        elif col_key == 'email':
            return obj.email or '-'
        elif col_key == 'phone':
            return obj.phone or '-'
        elif col_key == 'address':
            return obj.address or '-'
        elif col_key == 'is_active':
            return 'Activo' if obj.is_active else 'Inactivo'
        elif col_key == 'taxpayer_type':
            try:
                return obj.profile.get_taxpayer_type_display()
            except Exception:
                return '-'
        elif col_key == 'payment_terms':
            try:
                return obj.profile.get_payment_terms_display()
            except Exception:
                return '-'
        elif col_key == 'credit_limit':
            try:
                return f'{obj.profile.credit_limit:.2f}'
            except Exception:
                return '-'
        elif col_key == 'created_at':
            return obj.created_at.strftime('%d/%m/%Y %H:%M') if obj.created_at else '-'
        elif col_key == 'updated_at':
            return obj.updated_at.strftime('%d/%m/%Y %H:%M') if obj.updated_at else '-'
        else:
            return getattr(obj, col_key, '-')

    def export_excel_with_visible_columns(self):
        """Exportar a Excel con columnas visibles"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        visible_columns = self.get_customer_visible_cols()
        all_columns = get_all_customer_columns()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Clientes'[:31]

        headers = [all_columns[col]['label'] for col in visible_columns]
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='343A40')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for row_obj in self.get_queryset():
            row = [str(self.get_export_value(row_obj, col)) for col in visible_columns]
            ws.append(row)

        widths = [len(h) for h in headers]
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for c, cell in enumerate(row):
                cell.border = border
                widths[c] = max(widths[c], len(str(cell.value or '')))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w + 4, 60)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Clientes_{timezone.localtime().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def export_pdf_with_visible_columns(self):
        """Exportar a PDF con columnas visibles"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from django.http import HttpResponse

        visible_columns = self.get_customer_visible_cols()
        all_columns = get_all_customer_columns()

        response = HttpResponse(content_type='application/pdf')
        filename = f'Clientes_{timezone.localtime().strftime("%Y%m%d_%H%M")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        pagesize = landscape(A4) if len(visible_columns) > 5 else A4

        doc = SimpleDocTemplate(
            response, pagesize=pagesize,
            leftMargin=0.8 * cm, rightMargin=0.8 * cm,
            topMargin=0.8 * cm, bottomMargin=0.8 * cm,
        )

        styles = getSampleStyleSheet()

        if len(visible_columns) > 8:
            font_size = 7
        elif len(visible_columns) > 5:
            font_size = 8
        else:
            font_size = 9

        cell_style = ParagraphStyle('cell', parent=styles['Normal'],
                                   fontSize=font_size, leading=font_size + 2)
        head_style = ParagraphStyle('cellHead', parent=styles['Normal'],
                                   fontSize=font_size, leading=font_size + 2,
                                   textColor=colors.white, fontName='Helvetica-Bold')

        elements = [
            Paragraph('Listado de Clientes', styles['Title']),
            Paragraph(
                f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
                styles['Normal'],
            ),
            Spacer(1, 0.3 * cm),
        ]

        headers = [all_columns[col]['label'] for col in visible_columns]
        data = [headers]

        for obj in self.get_queryset():
            row = [str(self.get_export_value(obj, col)) for col in visible_columns]
            data.append(row)

        table = Table(data, colWidths=[(pagesize[0] - 1.6 * cm) / len(visible_columns)] * len(visible_columns))

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()

        visible_columns = self.request.session.get('customer_visible_columns', CUSTOMER_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns = validate_customer_visible_columns(visible_columns)

        ctx['visible_columns'] = get_customer_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_customer_columns()
        ctx['visible_columns_list'] = visible_columns

        return ctx

class CustomerCreateView(LoginRequiredMixin, CreateView):
    model = Customer; form_class = CustomerForm
    template_name = 'billing/customer_form.html'
    success_url = reverse_lazy('billing:customer_list')

class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    model = Customer; form_class = CustomerForm
    template_name = 'billing/customer_form.html'
    success_url = reverse_lazy('billing:customer_list')

class CustomerDetailView(LoginRequiredMixin, DetailView):
    model = Customer
    template_name = 'billing/customer_detail.html'
    context_object_name = 'customer'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Incluir perfil e historial de facturas del cliente
        customer = self.object
        ctx['invoices'] = customer.invoices.order_by('-invoice_date')[:10]
        try:
            ctx['profile'] = customer.profile
        except CustomerProfile.DoesNotExist:
            ctx['profile'] = None
        return ctx

class CustomerDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    model = Customer
    template_name = 'billing/customer_confirm_delete.html'
    success_url = reverse_lazy('billing:customer_list')
    staff_redirect_url = '/customers/'


@login_required
def customer_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de clientes"""
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        import json
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])

        is_valid, validated_columns = validate_customer_visible_columns(visible_columns)

        request.session['customer_visible_columns'] = validated_columns
        request.session.modified = True

        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_customer_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# === INVOICE (CBV) ===
class InvoiceListView(ExportListMixin, LoginRequiredMixin, ListView):
    model = Invoice
    template_name = 'billing/invoice_list.html'
    context_object_name = 'items'
    paginate_by = 10

    export_title = 'Facturas'
    export_fields = [
        ('#', 'id'),
        ('Cliente', lambda o: str(o.customer)),
        ('Fecha', lambda o: o.invoice_date.strftime('%d/%m/%Y %H:%M')),
        ('Subtotal', lambda o: f'{o.subtotal:.2f}'),
        ('Impuesto', lambda o: f'{o.tax:.2f}'),
        ('Total', lambda o: f'{o.total:.2f}'),
    ]

    def get(self, request, *args, **kwargs):
        """Manejar exportaciones con columnas visibles"""
        fmt = request.GET.get('export')
        if fmt == 'excel':
            return self.export_excel_with_visible_columns()
        if fmt == 'pdf':
            return self.export_pdf_with_visible_columns()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Invoice.objects.select_related('customer')
        g = self.request.GET
        if customer := g.get('customer', '').strip():
            qs = qs.filter(
                customer__first_name__icontains=customer
            ) | qs.filter(customer__last_name__icontains=customer)
        if date_from := g.get('date_from', '').strip():
            qs = qs.filter(invoice_date__date__gte=date_from)
        if date_to := g.get('date_to', '').strip():
            qs = qs.filter(invoice_date__date__lte=date_to)
        if total_min := g.get('total_min', '').strip():
            qs = qs.filter(total__gte=total_min)
        if total_max := g.get('total_max', '').strip():
            qs = qs.filter(total__lte=total_max)
        return qs

    def get_invoice_visible_cols(self):
        """Obtener columnas visibles de la sesión"""
        visible_columns_list = self.request.session.get('invoice_visible_columns', INVOICE_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns_list = validate_invoice_visible_columns(visible_columns_list)
        return visible_columns_list

    def get_export_value(self, obj, col_key):
        """Obtener el valor de una columna específica"""
        if col_key == 'id':
            return obj.id
        elif col_key == 'customer':
            return str(obj.customer)
        elif col_key == 'customer_dni':
            return obj.customer.dni if obj.customer else '-'
        elif col_key == 'invoice_date':
            return obj.invoice_date.strftime('%d/%m/%Y %H:%M') if obj.invoice_date else '-'
        elif col_key == 'num_items':
            return obj.details.count()
        elif col_key == 'subtotal':
            return f'{obj.subtotal:.2f}'
        elif col_key == 'tax':
            return f'{obj.tax:.2f}'
        elif col_key == 'total':
            return f'{obj.total:.2f}'
        elif col_key == 'is_active':
            return 'Activa' if obj.is_active else 'Anulada'
        else:
            return getattr(obj, col_key, '-')

    def export_excel_with_visible_columns(self):
        """Exportar a Excel con columnas visibles"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        visible_columns = self.get_invoice_visible_cols()
        all_columns = get_all_invoice_columns()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Facturas'[:31]

        headers = [all_columns[col]['label'] for col in visible_columns]
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='343A40')
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for row_obj in self.get_queryset():
            row = [str(self.get_export_value(row_obj, col)) for col in visible_columns]
            ws.append(row)

        widths = [len(h) for h in headers]
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for c, cell in enumerate(row):
                cell.border = border
                widths[c] = max(widths[c], len(str(cell.value or '')))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(w + 4, 60)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Facturas_{timezone.localtime().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def export_pdf_with_visible_columns(self):
        """Exportar a PDF con columnas visibles"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from django.http import HttpResponse

        visible_columns = self.get_invoice_visible_cols()
        all_columns = get_all_invoice_columns()

        response = HttpResponse(content_type='application/pdf')
        filename = f'Facturas_{timezone.localtime().strftime("%Y%m%d_%H%M")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        pagesize = landscape(A4) if len(visible_columns) > 5 else A4

        doc = SimpleDocTemplate(
            response, pagesize=pagesize,
            leftMargin=0.8 * cm, rightMargin=0.8 * cm,
            topMargin=0.8 * cm, bottomMargin=0.8 * cm,
        )

        styles = getSampleStyleSheet()

        if len(visible_columns) > 8:
            font_size = 7
        elif len(visible_columns) > 5:
            font_size = 8
        else:
            font_size = 9

        elements = [
            Paragraph('Listado de Facturas', styles['Title']),
            Paragraph(
                f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
                styles['Normal'],
            ),
            Spacer(1, 0.3 * cm),
        ]

        headers = [all_columns[col]['label'] for col in visible_columns]
        data = [headers]

        for obj in self.get_queryset():
            row = [str(self.get_export_value(obj, col)) for col in visible_columns]
            data.append(row)

        table = Table(data, colWidths=[(pagesize[0] - 1.6 * cm) / len(visible_columns)] * len(visible_columns))

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), font_size),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop('page', None)
        ctx['search_params'] = params.urlencode()

        visible_columns = self.request.session.get('invoice_visible_columns', INVOICE_DEFAULT_VISIBLE_COLUMNS)
        is_valid, visible_columns = validate_invoice_visible_columns(visible_columns)

        ctx['visible_columns'] = get_invoice_visible_columns(visible_columns)
        ctx['all_columns'] = get_all_invoice_columns()
        ctx['visible_columns_list'] = visible_columns

        return ctx

@login_required
def invoice_create(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        formset = InvoiceDetailFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            # Validar stock antes de guardar
            stock_errors = []
            for detail_form in formset:
                if detail_form.cleaned_data and not detail_form.cleaned_data.get('DELETE', False):
                    product = detail_form.cleaned_data.get('product')
                    quantity = detail_form.cleaned_data.get('quantity', 0)
                    if product and quantity:
                        if quantity > product.stock:
                            stock_errors.append(
                                f'"{product.name}" solo tiene {product.stock} unidades en stock, '
                                f'pero intentas vender {quantity}.'
                            )
            
            if stock_errors:
                for error in stock_errors:
                    messages.error(request, error)
            else:
                invoice = form.save()
                formset.instance = invoice
                formset.save()
                
                # Descontar stock de cada producto vendido
                for detail in invoice.details.all():
                    product = detail.product
                    product.stock -= detail.quantity
                    product.save()
                
                subtotal = sum(d.subtotal for d in invoice.details.all())
                invoice.subtotal = subtotal
                invoice.tax = subtotal * Decimal('0.15')
                invoice.total = invoice.subtotal + invoice.tax
                invoice.save()
                messages.success(request, f'Factura #{invoice.id} creada! Total: ${invoice.total}')
                return redirect('billing:invoice_list')
    else:
        form = InvoiceForm()
        formset = InvoiceDetailFormSet()
    return render(request, 'billing/invoice_form.html', {
        'form': form, 'formset': formset, 'title': 'Nueva Factura',
        'today': timezone.localtime().strftime('%d/%m/%Y'),
    })


@login_required
def api_product_info(request, pk):
    """API para obtener precio y stock de un producto (usado en el formulario de factura)."""
    from django.http import JsonResponse
    product = get_object_or_404(Product, pk=pk)
    return JsonResponse({
        'id': product.id,
        'name': product.name,
        'unit_price': str(product.unit_price),
        'stock': product.stock,
    })


class InvoiceDetailView(LoginRequiredMixin, DetailView):
    model = Invoice
    template_name = 'billing/invoice_detail.html'
    context_object_name = 'invoice'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        invoice = self.object
        details = invoice.details.select_related('product', 'product__brand').all()
        ctx['details'] = details
        ctx['total_units'] = sum(d.quantity for d in details)
        return ctx


@login_required
def invoice_update_visible_columns(request):
    """Actualizar columnas visibles para el listado de facturas"""
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        import json
        data = json.loads(request.body)
        visible_columns = data.get('visible_columns', [])

        is_valid, validated_columns = validate_invoice_visible_columns(visible_columns)

        request.session['invoice_visible_columns'] = validated_columns
        request.session.modified = True

        return JsonResponse({
            'success': True,
            'visible_columns': validated_columns,
            'message': f'Mostrando {len(validated_columns)} de {len(get_all_invoice_columns())} columnas'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


class InvoiceDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    model = Invoice
    template_name = 'billing/invoice_confirm_delete.html'
    success_url = reverse_lazy('billing:invoice_list')
    staff_redirect_url = '/invoices/'
